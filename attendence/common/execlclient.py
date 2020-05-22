# coding=utf-8

# from openpyxl import load_workbook  # do not support old xls file
from openpyxl import Workbook
from openpyxl.styles import  PatternFill
import xlrd
import logging
import os


logger = logging.getLogger('default')


class ExcelParse(object):
    """
    根据传入的key索引文件，解析execl文件，读取所有sheet
    默认读取所有字段；
    :return
    [ {
        key1:value,
        key2:value,
        ...
        },
        {
        key1:value,
        key2:value,
        ...
        },
        {...}
    ]
    """
    KEY_DEFAULT_NUM = 0
    SHEET_NUM = 0

    def __init__(self, execlfile=None, encoding='utf-8'):
        self.file = execlfile
        self._data = []
        if self.file:
            self.read_execl(self.file, encoding)
        self.workbook = Workbook()

    def read_execl(self, execlfile, encoding):
        try:
            workbook = xlrd.open_workbook(execlfile, encoding_override=encoding)
        except Exception as why:
            print why
        for sheetname in workbook.sheet_names():
            booksheet = workbook.sheet_by_name(sheetname)
            maxrows = booksheet.nrows
            maxcols = booksheet.ncols
            keys = booksheet.row_values(ExcelParse.KEY_DEFAULT_NUM)
            for item in range(maxrows-1):
                record = {}
                for index in range(maxcols):
                    record[keys[index]] = booksheet.row_values(item+1)[index]
                self._data.append(record)

    def write_execl(self, data=None):
        """写入execl"""
        booksheet = self.workbook.active
        for line in data:
            booksheet.append(line)

    def writeToMergeExecl(self, data, average_time, num):
        """考勤数据详情写入execl；
           并从第10列，第1行起，每隔7行合并一次单元格，并写入周平均出勤时间
        """
        booksheet = self.workbook.active
        booksheet.append(data[0])
        i = 0
        x = 9
        lastline = None
        for line in data[1:]:
            i += 1
            # new rules not use
            booksheet.append(line)
            booksheet = self.rules1(booksheet, lastline, line, i)
            lastline = line
            # end
            if i % num == 0:
                start = i - (num - 2)
                end = i + 1
                value = average_time.get(line[1], '0')
                booksheet.merge_cells(start_row=start, start_column=x, end_row=end, end_column=x, )
                booksheet.cell(start, x, value)

    def rules1(self, booksheet, lastline, line, row_num):
        """新加规则，10.30后打卡，必须满足前一天22.30下班 否则标红;
           此规则未使用"""

        def checktime(clocktime, kind):
            """判断是否符合时间设置
            """
            ruleminute = 30
            # 如果打卡时间是其他假期信息，忽略
            if len(clocktime.split(':')) <= 2:
                return True
            hour, minute, second = clocktime.split(':')
            if kind == 'clockout':
                # 下班打卡时间 满足大于 22:30 return True
                rulehour = 22
                if int(hour) > rulehour:
                    return True
                # 此处是判定是否隔天签退，凌晨签退
                if int(hour) <= 10:
                    return True
                if int(hour) == rulehour and int(minute) >= ruleminute:
                    return True
            if kind == 'clockin':
                # 上班打卡时间 满足小于10:30 return True
                rulehour = 10
                if int(hour) < rulehour:
                    return True
                if int(hour) == rulehour and int(minute) <= ruleminute:
                    return True
            return False

        # 如果获取不到打卡时间；忽略
        if len(line[4]) == 0:
            return booksheet
        # 打卡时间小于 10:30 忽略
        if checktime(line[4], 'clockin'):
            return booksheet
        # 如果有请假信息；忽略
        if len(line[7]) != 0:
            return booksheet

        # 第5列
        col = 'E'
        # 填充红色
        fill = PatternFill("solid", fgColor="FF3030")
        # 坐标
        location = col + str(row_num + 1)
        # 工号
        jobnumber = line[1]

        # 如果打卡时间大于10:30进行处理
        if not checktime(line[4], 'clockin'):
            # 上一条记录存在
            if lastline:
                last_jobnumber = lastline[1]
                # 工号相同，确保对比同一个人，而且签退记录不能为空
                if jobnumber == last_jobnumber and len(lastline[5])!= 0:
                    # 满足签退时间大于22:30,则符合条件，不予处理
                    if checktime(lastline[5], 'clockout'):
                        return booksheet

        # 其他情况默认标红
        cell = booksheet[location]
        cell.fill = fill
        return booksheet

    def save(self, filepath=None):
        """保存execl至指定路径"""
        try:
            logger.info('save execl_file  to %s', filepath)
            filepath = filepath or self.file
            self.workbook.save(filepath)
            # 写完后，清空当前工作表
            self.workbook = Workbook()
        except Exception as why:
            logger.error(why)

    @property
    def getExeclObj(self):
        return self.workbook

    @property
    def data(self):
        return self._data
